class LOL_opgg:
    import requests
    from lxml import etree
    import openpyxl
    def get_new_data_requests(self):
        headers={
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'
        }
        self.wb=self.openpyxl.Workbook()
        self.ws=self.wb.active
        self.count=1
        self.All_Hero=[]
        self.res=self.requests.get('http://www.op.gg/champion/statistics',headers=headers)
        self.html=self.etree.HTML(self.res.text)
        for self.tbody in self.html.xpath('//tbody'):
            self.trs=self.tbody.xpath('./tr')
            for self.tr in self.trs:
                self.td_top=self.tr.xpath('./td[1]/text()')[0]
                self.td_hero=self.tr.xpath('.//td[4]/a//div[@class="champion-index-table__name"]/text()')[0]
                self.td_hero_place=self.tr.xpath('.//td[4]/a/div[2]/text()')[0]
                self.td_win=self.tr.xpath('.//td[5]/text()')[0]
                self.td_race=self.tr.xpath('.//td[6]/text()')[0]
                self.ws['A'+str(self.count)]=self.td_hero
                self.ws['B'+str(self.count)]=self.td_top
                self.ws['C'+str(self.count)]=self.td_win
                self.ws['D'+str(self.count)]=self.td_race
                self.hero={self.td_hero:[self.td_top,self.td_win,self.td_race]}
                self.All_Hero.append(self.hero)
                self.count+=1
                #print(self.td_top)
        self.wb.save('LOL_Opgg_data.xlsx')
        print(self.All_Hero)
test=LOL_opgg()
test.get_new_data_requests()
